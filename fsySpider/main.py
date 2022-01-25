# -*- coding: utf-8 -*-

# @Time    : 2022/1/22 4:06
# @Author  : yeshenyong
# @File    : main.py

import os
import csv
import time
import json
from pyquery import PyQuery as pq
import pandas as pd
from selenium import webdriver
from openpyxl import load_workbook


class FinanceCrawler(object):
    def __init__(self, data, config):
        self.target_name = config["target_name"]
        self.config = config
        self.data = data
        self.company_list = config["company_list"][-1].split(',')

    def run(self):
        print(len(self.company_list))
        options = webdriver.ChromeOptions()
        # 添加头部
        options.add_argument(
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36'
        )
        driver = webdriver.Chrome("./chromedriver", chrome_options=options)
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {
                "source": """
                    Object.defineProperty(navigator, 'webdriver', {
                      get:() => undefined
                    })        
                """
            },
        )

        year_list = ['20' + ("%02d" % i) for i in range(20, 2, -1)]
        month_list = ['-12-31', '-09-30', '-06-30', '-03-31']
        year_list = [(year + month) for year in year_list for month in month_list]
        prof_map = {}
        debt_map = {}
        mone_map = {}
        prof_map['公司名'] = []
        debt_map['公司名'] = []
        mone_map['公司名'] = []

        # test_num = 40

        for company in self.company_list:
            try:
                driver.get(
                    f"http://basic.10jqka.com.cn/" + self.data[company] + "/finance.html#stockpage"
                )
            except:
                print(company, "超时")

            # print(driver.page_source)
            while "滑动验证码" in driver.page_source:
                print(company + "爬虫被识别")
                time.sleep(100)
                try:
                    driver.get(
                        f"http://basic.10jqka.com.cn/" + self.data[company] + "/finance.html#stockpage"
                    )
                except:
                    print(company + "超时")
            doc = pq(driver.page_source)

            """
            年份[]
            净利润[]
            资产负债比[]
            营业总收入[]
            """
            # kk 对应 map
            years_filter = list(set([i.text() for i in doc(f"tbody tr th div").items()]))
            years_filter.sort(reverse=True)
            data_filter = [i.text() for i in doc(f"tbody tr td").items()]
            data_title = [i.text() for i in doc(f"tbody tr th").items()]
            tag_index = self.get_tag_index(self.target_name, data_title)
            tag_index = tag_index[:len(self.target_name)]
            prof_list = []

            length = len(years_filter)
            print(tag_index)
            for idx in range(0, len(data_filter)):
                if idx % length == 0 and (idx/length) in tag_index:
                    prof_list.append([data_filter[idx + i] for i in range(length)])
            for idx in range(len(prof_list)):
                print(prof_list[idx])
            if len(prof_list) != 3:
                print(company + "有问题")
                continue
            prof_map['公司名'].append(company)
            debt_map['公司名'].append(company)
            mone_map['公司名'].append(company)

            for year in year_list:
                if year in years_filter:
                    try:
                        idx = years_filter.index(year)
                        if year not in prof_map:
                            prof_map[year] = []
                        if year not in debt_map:
                            debt_map[year] = []
                        if year not in mone_map:
                            mone_map[year] = []
                        prof_map[year].append(prof_list[0][idx])
                        debt_map[year].append(prof_list[1][idx])
                        mone_map[year].append(prof_list[2][idx])
                    except Exception as e:
                        print(company + "有问题")
                        break
                else:
                    if year not in prof_map:
                        prof_map[year] = []
                    if year not in debt_map:
                        debt_map[year] = []
                    if year not in mone_map:
                        mone_map[year] = []
                    prof_map[year].append("--")
                    debt_map[year].append("--")
                    mone_map[year].append("--")
            # break
            # test_num -= 1
            # if test_num < 0:
            #     break
            # ["中材股份, 宝胜科技, 中电控股, 东方航空, 中华国际, 金瑞科技, 中国中旅, 中国建材"]
            # time.sleep(random.randint(1, 4))
        self.parse_to_csv(year_list, prof_map, debt_map, mone_map)

    def get_tag_index(self, target_name, data_title):
        tag_index = []
        for i in range(len(data_title)):
            if data_title[i] in target_name:
                tag_index.append(i-1)
                # target_name.remove()
        return tag_index

    def calculate(self):
        """从company_list.json 中抽取想要的数据"""
        # 资产负债比率
        # 营业总收入(元)
        # 净利润

        self.parse_to_csv()  # 并将数据产生到csv 中

    def parse_to_csv(self, year_list, prof_map, debt_map, mone_map):
        # 资产负债比率
        # 营业总收入(元)
        # 净利润
        for index in prof_map:
            print(index, len(index), end="\t")
        dataframe = pd.DataFrame(prof_map)
        # 将DataFrame存储为csv,index表示是否显示行名，default=True
        dataframe.to_csv("净利润.csv", index=False, sep=',', encoding="utf_8_sig")
        dataframe = pd.DataFrame(debt_map)
        dataframe.to_csv("资产负债比率.csv", index=False, sep=',', encoding="utf_8_sig")
        dataframe = pd.DataFrame(mone_map)
        dataframe.to_csv("营业总收入.csv", index=False, sep=',', encoding="utf_8_sig")

        # self.reset_col_size(['净利润.csv', '资产负债比率.csv', '营业总收入.csv'], year_list)

    def reset_col_size(self, csv_list, year_list):
        for csv_item in csv_list:
            wb = load_workbook(csv_item)
            ws = wb[wb.sheetnames[0]]
            for year in year_list:
                ws.column_dimensions[year].width = 20.0
            wb.save(csv_item)

def loadData():
    data = {}
    config_file = os.path.join(os.getcwd(), 'config.json')
    file = os.path.join(os.getcwd(), 'map.txt')
    fd = open(file, 'r', encoding='gbk')
    with open(config_file, 'r', encoding='utf-8') as f:
        config = json.loads(f.read())
    while True:
        line = fd.readline()
        if not line:
            break
        line = line.strip().split(':')
        if len(line) != 2:
            continue
        # 中国黄金:000001
        data[line[0]] = line[1]
    fd.close()
    return data, config


if __name__ == '__main__':
    data, config = loadData()
    print(data)
    print(config)
    obj_spider = FinanceCrawler(data, config)
    obj_spider.run()